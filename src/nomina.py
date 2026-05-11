"""Genera el Excel de nómina mensual leyendo desde Google Sheets."""
import os

import gspread
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sheets import get_sheet

TECNICOS = [
    "CRISTIAN", "MARTIN", "ALVARO", "YOHAN", "ERCS",
    "HANS", "DIEGO", "LUIS E", "JEAN", "JOEL", "DIANA",
]
DESCUENTOS_FIJOS = [
    "ALTAS EN GARANTIA", "REUTILIZADAS GARANTIA",
    "AVERIAS EN GARANTIA", "IRPF", "EMBARGO",
]

_FILL_HEADER = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_FILL_ALT = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
_FONT_HEADER = Font(bold=True, name="Arial", size=10, color="FFFFFF")
_FONT_NORMAL = Font(name="Arial", size=10)
_FONT_BOLD = Font(bold=True, name="Arial", size=10)


def _limpiar_precio(v: str) -> float:
    return float(v.replace("€", "").replace(" ", "").replace(",", ".").strip() or "0")


def _leer_altas(wb: gspread.Spreadsheet, tecnico: str) -> list:
    try:
        rows = wb.worksheet(tecnico).get_all_values()
    except Exception:
        return []
    altas = []
    vistas: set[str] = set()
    for row in rows[2:]:
        if len(row) < 3 or not row[1] or not row[2]:
            continue
        orden, codigo = row[1].strip(), row[2].strip()
        if not orden or orden in ["-", "SECOMCOL"] or codigo in ["SIN ALTAS"] or "€" in codigo:
            continue
        if (orden, codigo) in vistas:
            continue
        vistas.add((orden, codigo))
        try:
            precio = _limpiar_precio(row[4]) if len(row) > 4 and row[4] else 0.0
        except ValueError:
            precio = 0.0
        altas.append([row[0].strip(), orden, codigo, precio])
    return altas


def _leer_descuentos(wb: gspread.Spreadsheet, tecnico: str) -> dict:
    try:
        rows = wb.worksheet("Descuentos").get_all_values()
    except Exception:
        return {}
    desc: dict[str, float] = {}
    for row in rows[1:]:
        if len(row) >= 4 and row[1].strip().upper() == tecnico.upper():
            concepto = row[2].strip().upper()
            try:
                monto = float(row[3].strip().replace(",", "."))
            except ValueError:
                monto = 0.0
            desc[concepto] = desc.get(concepto, 0.0) + monto
    return desc


def _escribir_hoja(ws, tecnico: str, altas: list, desc_vars: dict, mes: str, ano: str) -> None:
    COL_F, COL_O, COL_C, COL_P = 1, 2, 3, 4

    for col, w in [(COL_F, 12), (COL_O, 22), (COL_C, 12), (COL_P, 10)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    def _hcell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center")
        return c

    # Título
    _hcell(1, COL_F, f"{tecnico} — {mes.upper()} {ano}")
    ws.merge_cells(start_row=1, start_column=COL_F, end_row=1, end_column=COL_P)

    # Cabeceras
    for col, label in [(COL_F, "FECHA"), (COL_O, "ORDEN"), (COL_C, "CODIGO"), (COL_P, "TECNICO")]:
        _hcell(2, col, label)

    # Altas
    row = 3
    first_row = row
    for fecha, orden, codigo, precio in altas:
        for col, val in [(COL_F, fecha), (COL_O, orden), (COL_C, codigo), (COL_P, precio)]:
            ws.cell(row=row, column=col, value=val).font = _FONT_NORMAL
        row += 1
    last_row = row - 1

    # Festivos
    row += 1
    ws.cell(row=row, column=COL_O, value="FESTIVOS").font = _FONT_BOLD
    ws.cell(row=row, column=COL_P, value=0).font = _FONT_NORMAL
    festivos_row = row
    row += 1

    # Subtotal altas
    sub_row = row
    ws.cell(row=row, column=COL_P, value=f"=SUM(D{first_row}:D{last_row})+D{festivos_row}").font = _FONT_BOLD
    row += 2

    # Descuentos
    ws.cell(row=row, column=COL_O, value="DESCUENTOS").font = Font(bold=True, name="Arial", size=10, color="C00000")
    row += 1
    desc_start = row

    for concepto in DESCUENTOS_FIJOS:
        ws.cell(row=row, column=COL_O, value=concepto).font = _FONT_NORMAL
        ws.cell(row=row, column=COL_O).fill = _FILL_ALT
        ws.cell(row=row, column=COL_P, value=desc_vars.pop(concepto, 0)).font = _FONT_NORMAL
        ws.cell(row=row, column=COL_P).fill = _FILL_ALT
        row += 1

    for concepto, monto in sorted(desc_vars.items()):
        ws.cell(row=row, column=COL_O, value=concepto).font = _FONT_NORMAL
        ws.cell(row=row, column=COL_O).fill = _FILL_ALT
        ws.cell(row=row, column=COL_P, value=monto).font = _FONT_NORMAL
        ws.cell(row=row, column=COL_P).fill = _FILL_ALT
        row += 1

    desc_end = row - 1
    ws.cell(row=row, column=COL_O, value="SUBTOTAL").font = _FONT_BOLD
    ws.cell(row=row, column=COL_P, value=f"=SUM(D{desc_start}:D{desc_end})").font = _FONT_BOLD
    sub_desc_row = row
    row += 2

    # Total
    c = ws.cell(row=row, column=COL_O, value="TOTAL")
    c.font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
    c.fill = _FILL_HEADER
    ws.cell(row=row, column=COL_P, value=f"=D{sub_row}-D{sub_desc_row}").font = Font(bold=True, name="Arial", size=12)


def generar_excel(mes: str, ano: str) -> str:
    wb_sheet = get_sheet()
    wb = Workbook()
    wb.remove(wb.active)
    for tecnico in TECNICOS:
        altas = _leer_altas(wb_sheet, tecnico)
        desc = _leer_descuentos(wb_sheet, tecnico)
        _escribir_hoja(wb.create_sheet(title=tecnico), tecnico, altas, desc, mes, ano)
    nombre = f"ALTAS_{mes.upper()}_{ano}.xlsx"
    wb.save(nombre)
    return nombre
